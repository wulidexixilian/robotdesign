from model.m_math import mass_combine, tensor, tear_tensor
from openpyxl import load_workbook
import numpy as np


def masspara_combine(filename, sheetname, max_row):
    m = 0
    cm = np.array([0, 0, 0])
    it = tensor(*np.array([0, 0, 0, 0, 0, 0]))
    table = load_workbook(filename)
    sheet = table[sheetname]
    for idx_r in range(15, max_row):
        item = []
        for idx_c in range(2, 13):
            item.append(sheet.cell(row=idx_r + 1, column=idx_c + 1).value)
        m_new = item[0]
        cm_new = np.array(item[1:4]) * 1e-3
        it_new = tensor(*(np.array(item[4:-1]) * 1e-6))
        m, cm, it = mass_combine(m, m_new, cm, cm_new, it, it_new)
    it = tear_tensor(it)
    return m, cm, it


filename = 'C:/Users/pei.sun/Desktop/InWork_KR_Cybertech_gear_replacement_mass_parameters_V02.xlsx'
sheetname = ['Baseframe', 'Rotation_Column', 'Linkarm_short', 'Linkarm_long']
n_row = [23, 23, 23, 23]

for idx in range(4):
    m, cm, it = masspara_combine(filename, sheetname[idx], n_row[idx])
    np.set_printoptions(precision=2)
    print('{}:\nmass: {:.2f}\ncm: {}'.format(sheetname[idx], m, cm * 1e3))
    np.set_printoptions(precision=6)
    print('IT: {}'.format(it))



